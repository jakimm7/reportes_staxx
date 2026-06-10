import openpyxl as openpyxl
from copy import copy

CANAL = "canal_venta"
PRODUCTO = "producto"
TC = "tipo_cambio"
CANTIDAD = "cantidad"
IVA = "iva"
VALOR_FINAL = "valor_venta"
VALOR_NETO = "valor_neto"
IBB = "ibb"
CARGO_VENTA = "cargo_venta"
COSTO_ENVIO = "costo_envio"
COSTO_ADMIN = "costo_admin"
COMISION_VENDEDOR = "costo_vendedor"
INGRESO_NETO = "ingreso_neto"
COSTO_IMPORTACION = "costo_importacion"
GANACIA = "ganancia"
PORCENTAJE_GANACIA = "porcentaje_ganacia"
CARGADO_EN_SISTEMA = "cargado_en_sistema"

CANAL_FACTURA = "VENDEDOR"
PLATAFORMA_E_COMMERCE = "ML"
MEDIO_E_COMMERCE = "MP"
MEDIO_FACTURA_A = "Transferencia"
ESTANTERIA_200 = "200 KG"
ESTANTERIA_300 = "300 KG"

COSTO_800 = 65.58
COSTO_1200 = 87.17

ALICUOTA_IVA = 1.21
ALICUOTA_COSTO_ADMIN = 0.07
ALICUOTA_COMISION = 0.03

COLUMNA_FLETE = "F"
COLUMNA_CANTIDAD = "G"
COLUMNA_UNITARIO = "H"
COLUMNA_CARGO = "I"
COLUMNA_IMPUESTOS = "J"
COLUMNA_ENVIO = "M"
COLUMNA_NETO = "N"
COLUMNA_COSTO_ADMIN = "O"
COLUMNA_TC = "P"
COLUMNA_COSTO_IMPO = "Q"
COLUMNA_COSTO_TOTAL = "R"
COLUMNA_BENEFICIO = "S"

def procesar_data(datos_venta, tc_venta):
    datos_venta[CANTIDAD] = int(datos_venta[CANTIDAD])
    datos_venta[TC] = tc_venta
    datos_venta[VALOR_FINAL] = float(datos_venta[VALOR_FINAL])
    datos_venta[IBB] = float(datos_venta[IBB])
    datos_venta[CARGO_VENTA] = float(datos_venta[CARGO_VENTA])
    datos_venta[COSTO_ENVIO] = 0.00
    datos_venta[IVA] = datos_venta[VALOR_FINAL] - (datos_venta[VALOR_FINAL] / ALICUOTA_IVA)
    datos_venta[VALOR_NETO] = datos_venta[VALOR_FINAL] - datos_venta[IVA] - datos_venta[IBB] - datos_venta[CARGO_VENTA] - datos_venta[COSTO_ENVIO]
    datos_venta[COSTO_ADMIN] = 0.00
    datos_venta[COMISION_VENDEDOR] = 0.00

    if datos_venta[CANAL] == CANAL_FACTURA:
        datos_venta[COMISION_VENDEDOR] = datos_venta[VALOR_NETO] * ALICUOTA_COMISION
    
    datos_venta[INGRESO_NETO] = 0

    if datos_venta[PRODUCTO] == ESTANTERIA_200:
        datos_venta[COSTO_IMPORTACION] = (COSTO_800 * tc_venta) * datos_venta[CANTIDAD]
    else:
        datos_venta[COSTO_IMPORTACION] = (COSTO_1200 * tc_venta) * datos_venta[CANTIDAD]
    datos_venta[COSTO_ADMIN] = datos_venta[COSTO_IMPORTACION] * ALICUOTA_COSTO_ADMIN
    datos_venta[INGRESO_NETO] = datos_venta[VALOR_NETO] - datos_venta[COMISION_VENDEDOR] - datos_venta[COSTO_ADMIN]

    datos_venta[GANACIA] = datos_venta[INGRESO_NETO] - datos_venta[COSTO_IMPORTACION]
    datos_venta[PORCENTAJE_GANACIA] = datos_venta[GANACIA] / datos_venta[COSTO_IMPORTACION]
    datos_venta[CARGADO_EN_SISTEMA] = "NO"
    
    return dict_to_array(datos_venta)

def dict_to_array(datos):
    claves = list(datos.keys())
    nueva_fila = []
    for clave in claves:
        nueva_fila.append(datos[clave])
    return nueva_fila

def transcribir_excel(nueva_fila, archivo_excel):
    wb = openpyxl.load_workbook(archivo_excel)
    hoja = wb.active

    # nueva_fila.append(f"""={COLUMNA_FLETE}{proxima_fila}+{COLUMNA_CARGO}{proxima_fila}+{COLUMNA_IMPUESTOS}{proxima_fila}+{COLUMNA_ENVIO}{proxima_fila}+
    #                   {COLUMNA_COSTO_ADMIN}{proxima_fila}+{COLUMNA_COSTO_IMPO}{proxima_fila}*{COLUMNA_CANTIDAD}{proxima_fila}""")
    # nueva_fila.append(f"={COLUMNA_NETO}{proxima_fila}-{COLUMNA_COSTO_TOTAL}{proxima_fila}")
    # nueva_fila.append(f"={COLUMNA_BENEFICIO}{proxima_fila}/{COLUMNA_NETO}{proxima_fila}")
    hoja.append(nueva_fila)
    fila_actual = hoja.max_row
    fila_plantilla = hoja.max_row + 1

    for col in range(1, len(nueva_fila)+1):
        celda_origen = hoja.cell(row=fila_plantilla, column=col)
        celda_destino = hoja.cell(row=fila_actual, column=col)

        copiar_estilo(celda_origen, celda_destino)

    # tabla = ws.tables.get("Table2")
    
    # if tabla:
    #     rango_actual = tabla.ref
        
    #     col_inicio, fila_inicio, col_fin, fila_fin = openpyxl.utils.cell.range_boundaries(rango_actual)
    #     nueva_fila_fin = ws.max_row
        
    #     nuevo_rango = f"{openpyxl.utils.cell.get_column_letter(col_inicio)}{fila_inicio}:{openpyxl.utils.cell.get_column_letter(col_fin)}{nueva_fila_fin}"
        
    #     tabla.ref = nuevo_rango
    
    wb.save(archivo_excel)

def copiar_estilo(celda_origen, celda_destino):
    celda_destino.font = copy(celda_origen.font)
    celda_destino.border = copy(celda_origen.border)
    celda_destino.fill = copy(celda_origen.fill)
    celda_destino.number_format = copy(celda_origen.number_format)
    celda_destino.alignment = copy(celda_origen.alignment)
    celda_destino.protection = copy(celda_origen.protection)
