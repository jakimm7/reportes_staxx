import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator
from utils.utils import copiar_estilo, procesar_fecha, limpiar_numero

CANAL = "canal"
PRODUCTO = "producto"
CANTIDAD = "cantidad"
NUM_OP = "num_op"
NUM_FC = "num_fc"
FECHA = "fecha"
NOMBRE = "nombre_razon_social"
DATOS_FACTURACION = "datos_facturacion"
DNI = "dni_cuit"
FORMA_PAGO = "forma_pago"
PAGO = "pago"
TC = "tipo_cambio"
VALOR_VENTA = "valor_venta"
CARGO_VENTA = "cargo_venta"
COSTO_ENVIO = "costo_envio"
IBB = "ibb"
VALOR_NETO = "valor_neto"
IVA = "iva"
COSTO_IMPORTACION = "costo_importacion"
COSTO_ADMIN = "costo_admin"
COMISION_VENDEDOR = "costo_vendedor"

POS_CANAL = 0
POS_PRODUCTO = 1
POS_CANTIDAD = 2
POS_NUM_OP = 3
POS_NUM_FC = 4
POS_FECHA = 5
POS_NOMBRE = 6
POS_DATOS_FACTURACION = 7
POS_DNI = 8
POS_FORMA_PAGO = 9
POS_COMENTARIOS = 10
POS_PAGO = 11
POS_TC = 12
POS_VALOR_VENTA = 13
POS_CARGO_VENTA = 14
POS_IBB = 15
POS_IVA = 16
POS_VALOR_NETO = 17
POS_COSTO_ADMIN = 18
POS_COMISION = 19
POS_COSTO_IMPO = 20

TAM_ARR = 21

PLATAFORMA_E_COMMERCE = "ML"
ESTANTERIA_200 = "200 KG"
ESTANTERIA_300 = "300 KG"

COSTO_200 = 65.58
COSTO_300 = 87.17

ALICUOTA_IVA = 1.21
ALICUOTA_COSTO_ADMIN = 0.07
ALICUOTA_COMISION = 0.04

COLUMNAS_FORMULAS = 4

def procesar_data(datos_crudos, tc_venta):
    datos_procesados = [None] * TAM_ARR
    datos_procesados[POS_CANAL] = datos_crudos[CANAL]
    datos_procesados[POS_PRODUCTO] = datos_crudos[PRODUCTO]
    datos_procesados[POS_CANTIDAD] = limpiar_numero(datos_crudos[CANTIDAD])
    datos_procesados[POS_NUM_OP] = datos_crudos[NUM_OP]
    datos_procesados[POS_NUM_FC] = datos_crudos[NUM_FC]
    datos_procesados[POS_FECHA] = procesar_fecha(datos_crudos[FECHA])
    datos_procesados[POS_NOMBRE] = datos_crudos[NOMBRE]
    datos_procesados[POS_DATOS_FACTURACION] = datos_crudos[DATOS_FACTURACION]
    datos_procesados[POS_DNI] = datos_crudos[DNI]
    datos_procesados[POS_FORMA_PAGO] = datos_crudos[FORMA_PAGO]
    datos_procesados[POS_PAGO] = datos_crudos[PAGO]
    datos_procesados[POS_TC] = tc_venta
    datos_procesados[POS_VALOR_VENTA] = limpiar_numero(datos_crudos[VALOR_VENTA])
    datos_procesados[POS_CARGO_VENTA] = limpiar_numero(datos_crudos[CARGO_VENTA])
    datos_procesados[POS_IBB] = limpiar_numero(datos_crudos[IBB])

    if datos_procesados[POS_CANAL] == PLATAFORMA_E_COMMERCE:
        datos_procesados[POS_VALOR_NETO] = (datos_procesados[POS_VALOR_VENTA] / ALICUOTA_IVA) - datos_procesados[POS_IBB]
        datos_procesados[POS_IVA] = datos_procesados[POS_VALOR_VENTA] - datos_procesados[POS_VALOR_NETO]
    else:
        datos_procesados[POS_VALOR_NETO] = limpiar_numero(datos_crudos[VALOR_NETO])
        datos_procesados[POS_IVA] = limpiar_numero(datos_crudos[IVA])

    costo_impo = 0
    if datos_procesados[POS_PRODUCTO] == ESTANTERIA_200:
        costo_impo = COSTO_200 * datos_procesados[POS_TC]
    else:
        costo_impo = COSTO_300 * datos_procesados[POS_TC]
    
    datos_procesados[POS_COSTO_ADMIN] = costo_impo * ALICUOTA_COSTO_ADMIN
    datos_procesados[POS_COMISION] = datos_procesados[POS_VALOR_NETO] * ALICUOTA_COMISION
    datos_procesados[POS_COSTO_IMPO] = costo_impo

    return datos_procesados

def transcribir_excel(ruta_excel, datos_venta, mes, nombre_tabla):
    libro = openpyxl.load_workbook(ruta_excel)
    ws, tabla = None
    hojas_a_buscar = [libro[mes]] if mes else libro.worksheets

    for hoja_actual in hojas_a_buscar:
        if nombre_tabla in hoja_actual.tables:
            ws = hoja_actual
            tabla = hoja_actual.tables[nombre_tabla]
            break

    if tabla is None:
        raise ValueError(f"No se encontró la tabla '{nombre_tabla}' en el archivo.")

    rango_actual = tabla.ref
    col_inicio, fila_inicio, col_fin, fila_fin = openpyxl.utils.cell.range_boundaries(rango_actual)

    fila_anterior = fila_fin
    nueva_fila = fila_fin + 1
    col_inicio_formulas = col_fin - COLUMNAS_FORMULAS + 1

    for i, valor in enumerate(datos_venta):
        col = col_inicio + i
        if col >= col_inicio_formulas:
            break
        celda_origen = ws.cell(row=fila_anterior, column=col)
        celda_destino = ws.cell(row=nueva_fila, column=col)
        celda_destino.value = valor
        copiar_estilo(celda_origen, celda_destino)

    for col in range(col_inicio_formulas, col_fin + 1):
        letra = get_column_letter(col)
        celda_origen = ws.cell(row=fila_anterior, column=col)
        celda_destino = ws.cell(row=nueva_fila, column=col)
        copiar_estilo(celda_origen, celda_destino)

        if isinstance(celda_origen.value, str) and celda_origen.value.startswith("="):
            formula_traducida = Translator(
                celda_origen.value,
                origin=f"{letra}{fila_anterior}"
            ).translate_formula(f"{letra}{nueva_fila}")
            celda_destino.value = formula_traducida

    nuevo_rango = f"{get_column_letter(col_inicio)}{fila_inicio}:{get_column_letter(col_fin)}{nueva_fila}"
    tabla.ref = nuevo_rango

    libro.save(ruta_excel)
